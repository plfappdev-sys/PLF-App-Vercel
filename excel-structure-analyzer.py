#!/usr/bin/env python3
"""
Excel File Structure Analyzer
Purpose: Examine the structure of the PLF Contributions Excel file
Created: September 21, 2025
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def analyze_excel_structure(file_path):
    """Analyze the structure of the Excel file"""
    print(f"Analyzing Excel file: {file_path}")
    print("=" * 60)
    
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Get all sheet names
        sheet_names = workbook.sheetnames
        print(f"Total sheets found: {len(sheet_names)}")
        print("Sheet names:", sheet_names)
        print()
        
        # Analyze each sheet
        for sheet_name in sheet_names:
            print(f"📊 Analyzing sheet: {sheet_name}")
            print("-" * 40)
            
            sheet = workbook[sheet_name]
            
            # Get basic sheet info
            print(f"Dimensions: {sheet.dimensions}")
            print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
            
            # Read first few rows to understand structure
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                print(f"Columns found: {len(df.columns)}")
                print("Column names:", list(df.columns))
                
                # Show sample data
                if not df.empty:
                    print("Sample data:")
                    print(df.head(3).to_string(index=False))
                
            except Exception as e:
                print(f"Error reading sheet {sheet_name}: {e}")
            
            print()
        
        # Additional analysis for specific patterns
        print("🔍 Looking for contribution data patterns...")
        contribution_sheets = [name for name in sheet_names if any(keyword in name.lower() for keyword in ['contribution', 'month', '202', 'jan', 'feb', 'mar'])]
        member_sheets = [name for name in sheet_names if any(keyword in name.lower() for keyword in ['member', 'customer', 'client', 'info'])]
        
        print(f"Potential contribution sheets: {contribution_sheets}")
        print(f"Potential member info sheets: {member_sheets}")
        
        # Close workbook
        workbook.close()
        
        return sheet_names, contribution_sheets, member_sheets
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")
        return [], [], []

def get_detailed_sheet_info(file_path, sheet_name):
    """Get detailed information about a specific sheet"""
    print(f"\n📋 Detailed analysis of sheet: {sheet_name}")
    print("-" * 50)
    
    try:
        # Read the entire sheet
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        
        print(f"Total rows: {len(df)}")
        print(f"Total columns: {len(df.columns)}")
        print("\nColumn details:")
        for i, col in enumerate(df.columns):
            print(f"  {i+1}. {col} (dtype: {df[col].dtype})")
            # Show unique values for categorical columns with few unique values
            if df[col].dtype == 'object' and df[col].nunique() < 10:
                print(f"     Unique values: {df[col].unique()}")
        
        print(f"\nData types summary:")
        print(df.dtypes.value_counts())
        
        print(f"\nFirst 5 rows:")
        print(df.head().to_string(index=False))
        
        return df
        
    except Exception as e:
        print(f"Error reading sheet {sheet_name}: {e}")
        return None

def main():
    """Main function"""
    excel_file = "NewBusLogic/Peoples Liberator Fund Contributions 2025 App.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"Error: Excel file not found at {excel_file}")
        print("Please make sure the file exists in the NewBusLogic folder")
        return
    
    print("PLF Excel File Structure Analysis")
    print("=" * 60)
    
    # Analyze overall structure
    sheet_names, contribution_sheets, member_sheets = analyze_excel_structure(excel_file)
    
    # If we found potential contribution sheets, analyze the first one in detail
    if contribution_sheets:
        get_detailed_sheet_info(excel_file, contribution_sheets[0])
    
    # If we found potential member sheets, analyze the first one in detail
    if member_sheets:
        get_detailed_sheet_info(excel_file, member_sheets[0])
    
    print("\n" + "=" * 60)
    print("ANALYSIS COMPLETE")
    print("Next steps:")
    print("1. Review the sheet structures above")
    print("2. Identify which sheets contain member information")
    print("3. Identify which sheets contain contribution data")
    print("4. Create mapping from Excel columns to database schema")

if __name__ == "__main__":
    main()
