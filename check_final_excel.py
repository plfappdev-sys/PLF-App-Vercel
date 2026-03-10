#!/usr/bin/env python3
"""
Check the FINAL Excel file structure
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def check_excel_file(file_path):
    """Check the structure of the Excel file"""
    print(f"Checking Excel file: {file_path}")
    print("=" * 60)
    
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return
    
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Get all sheet names
        sheet_names = workbook.sheetnames
        print(f"Total sheets found: {len(sheet_names)}")
        print("Sheet names:", sheet_names)
        print()
        
        # Show first few rows of each sheet
        for sheet_name in sheet_names[:5]:  # Limit to first 5 sheets
            print(f"📊 Sheet: {sheet_name}")
            print("-" * 40)
            
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=3)
                print(f"Columns: {len(df.columns)}")
                print("Column names:", list(df.columns))
                if not df.empty:
                    print("First 2 rows:")
                    print(df.head(2).to_string(index=False))
                print()
            except Exception as e:
                print(f"Error reading sheet: {e}")
                print()
        
        workbook.close()
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")

if __name__ == "__main__":
    # Check the FINAL Excel file mentioned in the task
    final_excel = "NewBusLogic/Peoples Liberator Fund Contributions 30 June 2025 26-2-16 FINAL.xlsx"
    check_excel_file(final_excel)