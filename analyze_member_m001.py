#!/usr/bin/env python3
"""
Analyze Babotshedi Malibe (M001) balance calculation
Purpose: Understand how R32,191.61 balance was calculated
"""

import pandas as pd
import openpyxl
from datetime import datetime

def analyze_member_m001():
    """Analyze Babotshedi Malibe's financial data"""
    excel_file = "NewBusLogic/Peoples Liberator Fund Contributions 30 June 2025.xlsx"
    
    print("🔍 ANALYZING BABOTSHEDI MALIBE (M001) FINANCIAL DATA")
    print("=" * 60)
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_file)
        
        # Check all sheets for M001 data
        print("📊 Searching for M001 data across all sheets...")
        
        # Focus on the most recent financial year sheets
        recent_sheets = ['2024-2025', '2023-2024', '2022-2023']
        
        for sheet_name in recent_sheets:
            if sheet_name in workbook.sheetnames:
                print(f"\n📋 Analyzing sheet: {sheet_name}")
                print("-" * 40)
                
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                # Look for M001 in the Member column
                if 'Member' in df.columns:
                    m001_data = df[df['Member'].str.contains('M001|Babotshedi|Malibe', na=False, case=False)]
                    
                    if not m001_data.empty:
                        print(f"✅ Found M001 data in {sheet_name}")
                        print(f"Total rows found: {len(m001_data)}")
                        
                        # Display relevant columns
                        relevant_cols = [col for col in df.columns if any(keyword in str(col).lower() for keyword in 
                                                                         ['member', 'balance', 'contribution', 'interest', 'closing', 'share'])]
                        
                        if relevant_cols:
                            print("\nRelevant columns found:")
                            for col in relevant_cols:
                                print(f"  - {col}")
                            
                            # Show M001 data with relevant columns
                            display_cols = ['Member'] + relevant_cols
                            display_cols = [col for col in display_cols if col in df.columns]
                            
                            print(f"\nM001 data in {sheet_name}:")
                            print(m001_data[display_cols].to_string(index=False))
                            
                            # Look for closing balance or share value
                            balance_cols = [col for col in df.columns if 'closing' in str(col).lower() or 'share' in str(col).lower() or 'balance' in str(col).lower()]
                            if balance_cols:
                                for col in balance_cols:
                                    if col in m001_data.columns:
                                        balance_value = m001_data[col].iloc[0]
                                        if pd.notna(balance_value):
                                            print(f"\n💰 {col}: R{balance_value:,.2f}")
                        
                        break  # Found the data, no need to check other sheets
                    else:
                        print(f"❌ M001 not found in {sheet_name}")
                else:
                    print(f"❌ No 'Member' column found in {sheet_name}")
        
        # If not found in recent sheets, check all sheets
        if not m001_data.empty:
            print("\n🔍 Checking all sheets for M001...")
            for sheet_name in workbook.sheetnames:
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=5)
                    if 'Member' in df.columns:
                        full_df = pd.read_excel(excel_file, sheet_name=sheet_name)
                        m001_data = full_df[full_df['Member'].str.contains('M001|Babotshedi|Malibe', na=False, case=False)]
                        if not m001_data.empty:
                            print(f"✅ Found M001 in {sheet_name}")
                            break
                except:
                    continue
        
        # Analyze the 2024-2025 sheet in detail for M001
        print("\n" + "=" * 60)
        print("📈 DETAILED ANALYSIS OF 2024-2025 DATA")
        print("=" * 60)
        
        if '2024-2025' in workbook.sheetnames:
            df_2025 = pd.read_excel(excel_file, sheet_name='2024-2025')
            
            # Find M001
            m001_2025 = df_2025[df_2025['Member'].str.contains('M001|Babotshedi|Malibe', na=False, case=False)]
            
            if not m001_2025.empty:
                print("✅ Found M001 in 2024-2025 sheet")
                
                # Get all numeric columns for analysis
                numeric_cols = m001_2025.select_dtypes(include=['number']).columns.tolist()
                
                print(f"\n📊 Available numeric columns for analysis:")
                for col in numeric_cols:
                    value = m001_2025[col].iloc[0]
                    if pd.notna(value) and value != 0:
                        print(f"  - {col}: R{value:,.2f}")
                
                # Look for key financial metrics
                key_metrics = ['Closing Balance', 'Share Value', 'Total Contribution', 'Interest Earned']
                found_metrics = []
                
                for metric in key_metrics:
                    for col in df_2025.columns:
                        if metric.lower() in str(col).lower():
                            if col in m001_2025.columns:
                                value = m001_2025[col].iloc[0]
                                if pd.notna(value):
                                    found_metrics.append((col, value))
                
                if found_metrics:
                    print(f"\n💰 KEY FINANCIAL METRICS FOR M001 (2024-2025):")
                    for col, value in found_metrics:
                        print(f"  - {col}: R{value:,.2f}")
                
                # Check if we can find R32,191.61
                target_balance = 32191.61
                close_matches = []
                
                for col in numeric_cols:
                    value = m001_2025[col].iloc[0]
                    if pd.notna(value) and abs(value - target_balance) < 100:  # Within R100
                        close_matches.append((col, value))
                
                if close_matches:
                    print(f"\n🎯 CLOSE MATCHES TO R32,191.61:")
                    for col, value in close_matches:
                        print(f"  - {col}: R{value:,.2f} (difference: R{abs(value - target_balance):.2f})")
                else:
                    print(f"\n❌ No close matches found for R32,191.61")
                    
                    # Show all balances for reference
                    balance_cols = [col for col in df_2025.columns if 'balance' in str(col).lower()]
                    if balance_cols:
                        print(f"\n📈 All balance-related values:")
                        for col in balance_cols:
                            if col in m001_2025.columns:
                                value = m001_2025[col].iloc[0]
                                if pd.notna(value):
                                    print(f"  - {col}: R{value:,.2f}")
        
        # Check if there's a reconciliation sheet
        if 'Recon' in workbook.sheetnames:
            print("\n" + "=" * 60)
            print("📋 CHECKING RECONCILIATION SHEET")
            print("=" * 60)
            
            df_recon = pd.read_excel(excel_file, sheet_name='Recon')
            m001_recon = df_recon[df_recon['Member'].str.contains('M001|Babotshedi|Malibe', na=False, case=False)]
            
            if not m001_recon.empty:
                print("✅ Found M001 in reconciliation sheet")
                print(m001_recon.to_string(index=False))
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error analyzing Excel file: {e}")

def calculate_expected_balance():
    """Calculate what the balance should be based on business rules"""
    print("\n" + "=" * 60)
    print("🧮 EXPECTED BALANCE CALCULATION")
    print("=" * 60)
    
    # Based on PLF business rules:
    # - R200 monthly contributions
    # - 5.5% annual interest (compounded daily)
    # - Starting from July 2018
    
    print("Business Rules:")
    print("  - Monthly contribution: R200")
    print("  - Annual interest rate: 5.5%")
    print("  - Daily compounding")
    print("  - Period: July 2018 to June 2025 (7 years)")
    
    # Simple calculation
    months = 7 * 12  # 7 years
    total_contributions = months * 200
    
    # Approximate interest (simplified)
    # Average balance over time would be roughly half of total contributions
    avg_balance = total_contributions / 2
    annual_interest = avg_balance * 0.055
    total_interest = annual_interest * 7
    
    expected_balance = total_contributions + total_interest
    
    print(f"\n📊 Expected Balance Calculation:")
    print(f"  - Total contributions (84 months): R{total_contributions:,.2f}")
    print(f"  - Estimated interest earned: R{total_interest:,.2f}")
    print(f"  - Expected total balance: R{expected_balance:,.2f}")
    print(f"  - Target balance: R32,191.61")
    print(f"  - Difference: R{abs(expected_balance - 32191.61):,.2f}")

if __name__ == "__main__":
    analyze_member_m001()
    calculate_expected_balance()
